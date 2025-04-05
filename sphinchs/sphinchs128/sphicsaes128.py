import os
import time
import csv
import shutil
from Crypto.Cipher import AES
from Crypto.Random import get_random_bytes
import pyspx.sha2_256f as sphincs
from openpyxl import load_workbook, Workbook

# Generate SPHINCS+ key pair
def generate_sphincs_keypair():
    seed = get_random_bytes(96)  # Correct seed size for SHA-256
    pk, sk = sphincs.generate_keypair(seed)
    return pk, sk

# SPHINCS+ signing
def sphincs_sign(message, sk):
    signed_message = sphincs.sign(message, sk)
    return signed_message

# SPHINCS+ signature verification
def sphincs_verify(message, signed_message, pk):
    try:
        sphincs.verify(message, signed_message, pk)
        return True
    except Exception as e:
        print("Verification failed:", e)
        return False

# AES-128 encryption
def aes_encrypt(plaintext, key):
    cipher = AES.new(key, AES.MODE_EAX)
    nonce = cipher.nonce
    ciphertext, tag = cipher.encrypt_and_digest(plaintext)
    return nonce, ciphertext, tag

# AES-128 decryption
def aes_decrypt(nonce, ciphertext, tag, key):
    cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
    plaintext = cipher.decrypt_and_verify(ciphertext, tag)
    return plaintext

# Write data to file
def write_to_file(filepath, data):
    with open(filepath, 'wb') as file:
        file.write(data)

# Write results to CSV file
def write_to_csv(filepath, data):
    with open(filepath, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)

# Read rows from XLSX file
def read_rows_from_xlsx(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows

# Write rows to XLSX file
def write_rows_to_xlsx(filepath, rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filepath)

# Compare two XLSX files
def compare_xlsx_files(file1, file2):
    rows1 = read_rows_from_xlsx(file1)
    rows2 = read_rows_from_xlsx(file2)
    return rows1 == rows2

# Create directories if they don't exist
def create_directories():
    directories = ['keys', 'encrypted', 'signed', 'decrypted', 'results']
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)

# Example usage
if __name__ == "__main__":
    create_directories()

    # SPHINCS+ key generation
    pk_sphincs, sk_sphincs = generate_sphincs_keypair()

    # Save SPHINCS+ public and private keys to files
    write_to_file('keys/public_key_sphincs.bin', pk_sphincs)
    write_to_file('keys/private_key_sphincs.bin', sk_sphincs)
    print("SPHINCS+ public and private keys written to keys/public_key_sphincs.bin and keys/private_key_sphincs.bin")

    # Read rows from XLSX file
    rows = read_rows_from_xlsx('city.xlsx')

    # AES-128 key generation
    aes_key = get_random_bytes(16)  # AES-128 key
    write_to_file('keys/aes_key.bin', aes_key)
    print("AES-128 key written to keys/aes_key.bin")

    # Initialize CSV file with headers
    write_to_csv('results/computation_times.csv', ['Iteration', 'Data Size (bytes)', 'Encrypt Time (seconds)', 'Sign Time (seconds)', 'Verify Time (seconds)', 'Decrypt Time (seconds)', 'Total Time (seconds)'])

    decrypted_rows = []

    # Perform the operations and measure time for each row
    for i, row in enumerate(rows):
        message = str(row).encode()  # Convert row to bytes
        data_size = len(message)

        start_time = time.time()

        # AES-128 encryption
        encrypt_start_time = time.time()
        nonce, ciphertext, tag = aes_encrypt(message, aes_key)
        encrypt_end_time = time.time()
        encrypt_time = encrypt_end_time - encrypt_start_time

        # SPHINCS+ signing
        sign_start_time = time.time()
        signed_message = sphincs_sign(ciphertext, sk_sphincs)
        sign_end_time = time.time()
        sign_time = sign_end_time - sign_start_time

        # SPHINCS+ verification
        verify_start_time = time.time()
        if not sphincs_verify(ciphertext, signed_message, pk_sphincs):
            print(f"Iteration {i}: SPHINCS+ verification failed")
            continue
        verify_end_time = time.time()
        verify_time = verify_end_time - verify_start_time

        # AES-128 decryption
        decrypt_start_time = time.time()
        decrypted_message = aes_decrypt(nonce, ciphertext, tag, aes_key)
        decrypt_end_time = time.time()
        decrypt_time = decrypt_end_time - decrypt_start_time

        # Verify the content of the original and decrypted message
        if message != decrypted_message:
            print(f"Iteration {i}: AES-128 decryption failed")
            continue

        end_time = time.time()
        total_time = end_time - start_time

        # Log data size and computation times to CSV
        write_to_csv('results/computation_times.csv', [i, data_size, encrypt_time, sign_time, verify_time, decrypt_time, total_time])
        print(f"Iteration {i}, Data Size: {data_size} bytes, Encrypt Time: {encrypt_time:.6f} seconds, Sign Time: {sign_time:.6f} seconds, Verify Time: {verify_time:.6f} seconds, Decrypt Time: {decrypt_time:.6f} seconds, Total Time: {total_time:.6f} seconds")

        # Save the encrypted, signed, and decrypted messages for each row
        write_to_file(f'encrypted/ciphertext_{i}.bin', nonce + ciphertext + tag)
        write_to_file(f'signed/signed_message_{i}.bin', signed_message)
        write_to_file(f'decrypted/decrypted_row_{i}.bin', decrypted_message)
        print(f"Row {i}: AES-128 ciphertext, signed message, and decrypted message written to respective files")

        # Append decrypted row to list
        decrypted_rows.append(row)

    # Write decrypted rows to a new XLSX file
    write_rows_to_xlsx('decrypted/decrypted_city.xlsx', decrypted_rows)
    print("Decrypted rows written to decrypted/decrypted_city.xlsx")

    # Compare the original and new XLSX files
    if compare_xlsx_files('city.xlsx', 'decrypted/decrypted_city.xlsx'):
        print("The original and decrypted XLSX files are the same.")
    else:
        print("The original and decrypted XLSX files are different.")
