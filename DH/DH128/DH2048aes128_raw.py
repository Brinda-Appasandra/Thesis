import os
import time
import csv
import shutil
import binascii
from Crypto.Cipher import AES
from Crypto.Hash import SHA256, HMAC
from Crypto.Random import get_random_bytes
from Crypto.Util.number import getPrime, inverse, GCD
from openpyxl import load_workbook, Workbook

# Manual Diffie-Hellman implementation
class DiffieHellman:
    def __init__(self, group_size=2048):
        # Use standard parameters for DH
        # These are IETF parameters for a 2048-bit DH group (Group 14)
        self.p = int('FFFFFFFFFFFFFFFFC90FDAA22168C234C4C6628B80DC1CD129024E088A67CC74020BBEA63B139B22514A08798E3404DDEF9519B3CD3A431B302B0A6DF25F14374FE1356D6D51C245E485B576625E7EC6F44C42E9A637ED6B0BFF5CB6F406B7EDEE386BFB5A899FA5AE9F24117C4B1FE649286651ECE45B3DC2007CB8A163BF0598DA48361C55D39A69163FA8FD24CF5F83655D23DCA3AD961C62F356208552BB9ED529077096966D670C354E4ABC9804F1746C08CA18217C32905E462E36CE3BE39E772C180E86039B2783A2EC07A28FB5C55DF06F4C52C9DE2BCBF6955817183995497CEA956AE515D2261898FA051015728E5A8AACAA68FFFFFFFFFFFFFFFF', 16)
        self.g = 2  # Generator
        self.private_key = None
        self.public_key = None
        self.shared_key = None
        
    def generate_key(self):
        # Generate a random private key (a)
        # For a 2048-bit prime, use a ~256-bit private key
        private_key_size = 256 // 8  # 32 bytes
        self.private_key = int.from_bytes(get_random_bytes(private_key_size), byteorder='big')
        
        # Calculate public key g^a mod p
        self.public_key = pow(self.g, self.private_key, self.p)
        return self.public_key.to_bytes((self.p.bit_length() + 7) // 8, byteorder='big')
    
    def generate_shared_key(self, other_public_key):
        # Convert bytes to integer if necessary
        if isinstance(other_public_key, bytes):
            other_public_key = int.from_bytes(other_public_key, byteorder='big')
            
        # Calculate shared secret (other_public_key^private_key mod p)
        self.shared_key = pow(other_public_key, self.private_key, self.p)
        shared_key_bytes = self.shared_key.to_bytes((self.p.bit_length() + 7) // 8, byteorder='big')
        
        # Use SHA-256 to derive a 32-byte key from the shared secret
        h = SHA256.new(shared_key_bytes)
        return h.digest()

# AES-256 encryption using EAX mode
def aes_encrypt(plaintext, key):
    cipher = AES.new(key, AES.MODE_EAX)
    nonce = cipher.nonce
    ciphertext, tag = cipher.encrypt_and_digest(plaintext)
    return nonce, ciphertext, tag

# AES-256 decryption using EAX mode
def aes_decrypt(nonce, ciphertext, tag, key):
    cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
    plaintext = cipher.decrypt_and_verify(ciphertext, tag)
    return plaintext

# HMAC signing
def hmac_sign(message, key):
    h = HMAC.new(key, message, SHA256)
    return h.digest()

# HMAC verification
def hmac_verify(message, signature, key):
    h = HMAC.new(key, message, SHA256)
    try:
        h.verify(signature)
        return True
    except ValueError:
        return False

# Write data to file
def write_to_file(filepath, data):
    with open(filepath, 'wb') as file:
        file.write(data)

# Read data from file
def read_from_file(filepath):
    with open(filepath, 'rb') as file:
        return file.read()

# Write results to CSV file
def write_to_csv(filepath, data):
    with open(filepath, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)

# Read rows from XLSX file
def read_rows_from_xlsx(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows

# Write rows to XLSX file
def write_rows_to_xlsx(filepath, rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filepath)

# Compare two XLSX files
def compare_xlsx_files(file1, file2):
    rows1 = read_rows_from_xlsx(file1)
    rows2 = read_rows_from_xlsx(file2)
    return rows1 == rows2

# Create directories if they don't exist
def create_directories():
    directories = ['keys', 'encrypted', 'decrypted', 'signed', 'results']
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)

# Example usage
if __name__ == "__main__":
    create_directories()

    # Generate DH key pairs (simulating two parties)
    alice_dh = DiffieHellman()
    bob_dh = DiffieHellman()
    
    alice_public = alice_dh.generate_key()
    bob_public = bob_dh.generate_key()
    
    # Save public keys
    write_to_file('keys/alice_public_key.bin', alice_public)
    write_to_file('keys/bob_public_key.bin', bob_public)
    print("DH public keys written to keys/alice_public_key.bin and keys/bob_public_key.bin")
    
    # Create shared secret
    shared_key_alice = alice_dh.generate_shared_key(bob_public)
    shared_key_bob = bob_dh.generate_shared_key(alice_public)
    
    # Verify both parties have the same shared secret
    if shared_key_alice == shared_key_bob:
        print("DH key exchange successful: shared secrets match")
        print(f"Shared key (first 10 bytes): {binascii.hexlify(shared_key_alice[:10]).decode()}")
    else:
        print("DH key exchange failed: shared secrets do not match")
        print(f"Alice's shared key: {binascii.hexlify(shared_key_alice[:10]).decode()}")
        print(f"Bob's shared key: {binascii.hexlify(shared_key_bob[:10]).decode()}")
        exit(1)
        
    # Use the shared secret as the AES key - for AES-128 we need 16 bytes
    aes_key = shared_key_alice[:16]  # Use first 16 bytes for AES-128
    write_to_file('keys/aes_key.bin', aes_key)
    print("AES-128 key derived from DH shared secret and saved to keys/aes_key.bin")
    
    # Create a separate HMAC key derived from the shared secret
    hmac_key = SHA256.new(shared_key_alice + b"HMAC_KEY").digest()
    write_to_file('keys/hmac_key.bin', hmac_key)
    print("HMAC key derived from DH shared secret and saved to keys/hmac_key.bin")

    # Read rows from XLSX file
    rows = read_rows_from_xlsx('city.xlsx')

    # Initialize CSV file with headers
    write_to_csv('results/computation_times.csv', ['Iteration', 'Data Size (bytes)', 'Encrypt Time (seconds)', 'Sign Time (seconds)', 'Verify Time (seconds)', 'Decrypt Time (seconds)', 'Total Time (seconds)'])

    decrypted_rows = []

    # Perform the operations and measure time for each row
    for i, row in enumerate(rows):
        message = str(row).encode()  # Convert row to bytes
        data_size = len(message)

        start_time = time.time()

        # AES encryption
        encrypt_start_time = time.time()
        nonce, ciphertext, tag = aes_encrypt(message, aes_key)
        encrypt_end_time = time.time()
        encrypt_time = encrypt_end_time - encrypt_start_time
        write_to_file(f'encrypted/ciphertext_{i}.bin', nonce + ciphertext + tag)
        print(f"Iteration {i}: AES-encrypted message saved to 'encrypted/ciphertext_{i}.bin'.")

        # HMAC signing (replaced RSA signing)
        sign_start_time = time.time()
        signature = hmac_sign(ciphertext, hmac_key)
        sign_end_time = time.time()
        sign_time = sign_end_time - sign_start_time
        write_to_file(f'signed/signature_{i}.bin', signature)
        print(f"Iteration {i}: HMAC signature saved to 'signed/signature_{i}.bin'.")

        # HMAC verification (replaced RSA verification)
        verify_start_time = time.time()
        signature_read = read_from_file(f'signed/signature_{i}.bin')
        is_verified = hmac_verify(ciphertext, signature_read, hmac_key)
        verify_end_time = time.time()
        verify_time = verify_end_time - verify_start_time
        print(f"Iteration {i}: Verification {'succeeded' if is_verified else 'failed'} for 'signed/signature_{i}.bin'.")

        # AES decryption
        decrypt_start_time = time.time()
        aes_ciphertext_data = read_from_file(f'encrypted/ciphertext_{i}.bin')
        nonce_read = aes_ciphertext_data[:16]
        tag_read = aes_ciphertext_data[-16:]
        ciphertext_read = aes_ciphertext_data[16:-16]
        decrypted_message = aes_decrypt(nonce_read, ciphertext_read, tag_read, aes_key)
        decrypt_end_time = time.time()
        decrypt_time = decrypt_end_time - decrypt_start_time
        write_to_file(f'decrypted/decrypted_row_{i}.bin', decrypted_message)
        print(f"Iteration {i}: Decrypted message saved to 'decrypted/decrypted_row_{i}.bin'.")

        end_time = time.time()
        total_time = end_time - start_time

        # Log data size and computation times to CSV
        write_to_csv('results/computation_times.csv', [i, data_size, encrypt_time, sign_time, verify_time, decrypt_time, total_time])
        print(f"Iteration {i}, Data Size: {data_size} bytes, Encrypt Time: {encrypt_time:.6f} seconds, Sign Time: {sign_time:.6f} seconds, Verify Time: {verify_time:.6f} seconds, Decrypt Time: {decrypt_time:.6f} seconds, Total Time: {total_time:.6f} seconds")

        # Verify the content of the original and decrypted message
        if message != decrypted_message:
            print(f"Iteration {i}: AES decryption failed")
            continue

        # Append decrypted row to list
        decrypted_rows.append(row)

    # Write decrypted rows to a new XLSX file
    write_rows_to_xlsx('decrypted/decrypted_city.xlsx', decrypted_rows)
    print("Decrypted rows written to decrypted/decrypted_city.xlsx")

    # Compare the original and new XLSX files
    if compare_xlsx_files('city.xlsx', 'decrypted/decrypted_city.xlsx'):
        print("The original and decrypted XLSX files are the same.")
    else:
        print("The original and decrypted XLSX files are different.")
