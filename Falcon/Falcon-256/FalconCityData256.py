import os
import time
import csv
import shutil
from Crypto.Cipher import AES
from Crypto.Hash import SHA256
from Crypto.Random import get_random_bytes
from openpyxl import load_workbook, Workbook
import oqs

# Generate Falcon key pair using python-oqs
def generate_falcon_keypair():
    with oqs.Signature("Falcon-512") as sig:
        public_key = sig.generate_keypair()
        secret_key = sig.export_secret_key()
    return public_key, secret_key

# Falcon signing: generate signature using the secret key
def falcon_sign(message, secret_key):
    with oqs.Signature("Falcon-512", secret_key=secret_key) as sig:
        signature = sig.sign(message)
    return signature

# Falcon verification: verify signature using the public key
def falcon_verify(message, signature, public_key):
    with oqs.Signature("Falcon-512") as sig:
        return sig.verify(message, signature, public_key)

# Derive a 256-bit AES key from a password using SHA-256 (32 bytes)
def derive_aes_key(password):
    hash_obj = SHA256.new(password.encode())
    return hash_obj.digest()  # AES-256 key will be 32 bytes

# AES-256 encryption using EAX mode
def aes_encrypt(plaintext, key):
    cipher = AES.new(key, AES.MODE_EAX)
    nonce = cipher.nonce
    ciphertext, tag = cipher.encrypt_and_digest(plaintext)
    return nonce, ciphertext, tag

# AES-256 decryption using EAX mode
def aes_decrypt(nonce, ciphertext, tag, key):
    cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
    plaintext = cipher.decrypt_and_verify(ciphertext, tag)
    return plaintext

# Write data to file
def write_to_file(filepath, data):
    with open(filepath, 'wb') as file:
        file.write(data)

# Read data from file
def read_from_file(filepath):
    with open(filepath, 'rb') as file:
        return file.read()

# Write results to CSV file
def write_to_csv(filepath, data):
    with open(filepath, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)

# Read rows from XLSX file
def read_rows_from_xlsx(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(row)
    return rows

# Write rows to XLSX file
def write_rows_to_xlsx(filepath, rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filepath)

# Compare two XLSX files
def compare_xlsx_files(file1, file2):
    rows1 = read_rows_from_xlsx(file1)
    rows2 = read_rows_from_xlsx(file2)
    return rows1 == rows2

# Create directories if they don't exist
def create_directories():
    directories = ['keys', 'encrypted', 'decrypted', 'signed', 'results']
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)

# Example usage
if __name__ == "__main__":
    create_directories()

    # Generate Falcon key pair and save keys
    public_key, secret_key = generate_falcon_keypair()
    write_to_file('keys/public_key_falcon.bin', public_key)
    write_to_file('keys/private_key_falcon.bin', secret_key)
    print("Falcon public and private keys written to keys/public_key_falcon.bin and keys/private_key_falcon.bin")

    # Derive AES-256 key from a password and save it
    aes_key = derive_aes_key("password")  # AES-256 key (32 bytes)
    write_to_file('keys/aes_key.bin', aes_key)
    print("AES-256 key derived and saved to keys/aes_key.bin")

    # Read rows from XLSX file
    rows = read_rows_from_xlsx('city.xlsx')

    # Initialize CSV file with headers
    write_to_csv('results/computation_times.csv', ['Iteration', 'Data Size (bytes)', 'Encrypt Time (seconds)', 'Sign Time (seconds)', 'Verify Time (seconds)', 'Decrypt Time (seconds)', 'Total Time (seconds)'])

    decrypted_rows = []

    # Perform the operations and measure time for each row
    for i, row in enumerate(rows):
        message = str(row).encode()  # Convert row to bytes
        data_size = len(message)

        start_time = time.time()

        # AES-256 encryption
        encrypt_start_time = time.time()
        nonce, ciphertext, tag = aes_encrypt(message, aes_key)
        encrypt_end_time = time.time()
        encrypt_time = encrypt_end_time - encrypt_start_time
        write_to_file(f'encrypted/ciphertext_{i}.bin', nonce + ciphertext + tag)
        print(f"Iteration {i}: AES-encrypted message saved to 'encrypted/ciphertext_{i}.bin'.")

        # Falcon signing
        sign_start_time = time.time()
        signature = falcon_sign(ciphertext, secret_key)
        sign_end_time = time.time()
        sign_time = sign_end_time - sign_start_time
        write_to_file(f'signed/signature_{i}.bin', signature)
        print(f"Iteration {i}: Falcon signature saved to 'signed/signature_{i}.bin'.")

        # Falcon verification
        verify_start_time = time.time()
        signature_read = read_from_file(f'signed/signature_{i}.bin')
        is_verified = falcon_verify(ciphertext, signature_read, public_key)
        verify_end_time = time.time()
        verify_time = verify_end_time - verify_start_time
        print(f"Iteration {i}: Verification {'succeeded' if is_verified else 'failed'} for 'signed/signature_{i}.bin'.")

        # AES-256 decryption
        decrypt_start_time = time.time()
        aes_ciphertext_data = read_from_file(f'encrypted/ciphertext_{i}.bin')
        nonce_read = aes_ciphertext_data[:16]
        tag_read = aes_ciphertext_data[-16:]
        ciphertext_read = aes_ciphertext_data[16:-16]
        decrypted_message = aes_decrypt(nonce_read, ciphertext_read, tag_read, aes_key)
        decrypt_end_time = time.time()
        decrypt_time = decrypt_end_time - decrypt_start_time
        write_to_file(f'decrypted/decrypted_row_{i}.bin', decrypted_message)
        print(f"Iteration {i}: Decrypted message saved to 'decrypted/decrypted_row_{i}.bin'.")

        end_time = time.time()
        total_time = end_time - start_time

        # Log data size and computation times to CSV
        write_to_csv('results/computation_times.csv', [i, data_size, encrypt_time, sign_time, verify_time, decrypt_time, total_time])
        print(f"Iteration {i}, Data Size: {data_size} bytes, Encrypt Time: {encrypt_time:.6f} seconds, Sign Time: {sign_time:.6f} seconds, Verify Time: {verify_time:.6f} seconds, Decrypt Time: {decrypt_time:.6f} seconds, Total Time: {total_time:.6f} seconds")

        # Verify the content of the original and decrypted message
        if message != decrypted_message:
            print(f"Iteration {i}: AES-256 decryption failed")
            continue

        # Append decrypted row to list
        decrypted_rows.append(row)

    # Write decrypted rows to a new XLSX file
    write_rows_to_xlsx('decrypted/decrypted_city.xlsx', decrypted_rows)
    print("Decrypted rows written to decrypted/decrypted_city.xlsx")

    # Compare the original and new XLSX files
    if compare_xlsx_files('city.xlsx', 'decrypted/decrypted_city.xlsx'):
        print("The original and decrypted XLSX files are the same.")
    else:
        print("The original and decrypted XLSX files are different.")
