import os
import time
import csv
import shutil
from Crypto.Cipher import AES
from openpyxl import load_workbook, Workbook
import oqs

# AES-256 encryption using EAX mode
def aes_encrypt(plaintext, key):
    cipher = AES.new(key, AES.MODE_EAX)
    nonce = cipher.nonce
    ciphertext, tag = cipher.encrypt_and_digest(plaintext)
    return nonce, ciphertext, tag

# AES-256 decryption using EAX mode
def aes_decrypt(nonce, ciphertext, tag, key):
    cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
    plaintext = cipher.decrypt_and_verify(ciphertext, tag)
    return plaintext

# Write binary data to a file
def write_to_file(filepath, data):
    with open(filepath, 'wb') as file:
        file.write(data)

# Read binary data from a file
def read_from_file(filepath):
    with open(filepath, 'rb') as file:
        return file.read()

# Write a row to CSV
def write_to_csv(filepath, data):
    with open(filepath, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)

# Read rows from an XLSX file
def read_rows_from_xlsx(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    return [row for row in sheet.iter_rows(values_only=True)]

# Write rows to an XLSX file
def write_rows_to_xlsx(filepath, rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filepath)

# Compare two XLSX files
def compare_xlsx_files(file1, file2):
    return read_rows_from_xlsx(file1) == read_rows_from_xlsx(file2)

# Create required directories if they don't exist
def create_directories():
    for d in ['keys', 'encrypted', 'decrypted', 'results']:
        if os.path.exists(d):
            shutil.rmtree(d)
        os.makedirs(d)

# MAIN
if __name__ == "__main__":
    create_directories()

    # Generate Kyber512 KEM key pair
    with oqs.KeyEncapsulation("Kyber512") as kem:
        public_key = kem.generate_keypair()
        private_key = kem.export_secret_key()
        write_to_file("keys/public_key_kyber.bin", public_key)
        write_to_file("keys/private_key_kyber.bin", private_key)
        print("Kyber keypair generated and saved.")

    # Read rows from XLSX file
    rows = read_rows_from_xlsx("city.xlsx")

    # Initialize results CSV file with headers
    write_to_csv('results/computation_times.csv', [
        'Iteration', 'Data Size (bytes)', 'Encapsulation Time (s)',
        'Encrypt Time (s)', 'Decrypt Time (s)', 'Decapsulation Time (s)', 'Total Time (s)'
    ])

    decrypted_rows = []

    for i, row in enumerate(rows):
        message = str(row).encode()  # Convert row to bytes
        data_size = len(message)
        start_time = time.time()

        # Encapsulate shared secret using Kyber512 by passing the public key as an argument
        with oqs.KeyEncapsulation("Kyber512") as kem_enc:
            encap_start = time.time()
            ciphertext_kem, shared_secret_enc = kem_enc.encap_secret(public_key)
            encap_end = time.time()
            encap_time = encap_end - encap_start

        # Derive AES-256 key from the shared secret (expected to be 32 bytes)
        aes_key = shared_secret_enc[:32]

        # Save the derived AES-256 key for record
        write_to_file("keys/aes_key.bin", aes_key)
        print("AES-256 key derived and saved to keys/aes_key.bin.")

        # AES-256 encryption
        encrypt_start = time.time()
        nonce, ciphertext, tag = aes_encrypt(message, aes_key)
        encrypt_end = time.time()
        encrypt_time = encrypt_end - encrypt_start

        write_to_file(f'encrypted/ciphertext_{i}.bin', nonce + ciphertext + tag)
        write_to_file(f'encrypted/kem_ciphertext_{i}.bin', ciphertext_kem)

        # Decapsulate shared secret using Kyber512 by providing the private key in the constructor
        with oqs.KeyEncapsulation("Kyber512", secret_key=private_key) as kem_dec:
            decap_start = time.time()
            shared_secret_dec = kem_dec.decap_secret(ciphertext_kem)
            decap_end = time.time()
            decap_time = decap_end - decap_start

        aes_key_dec = shared_secret_dec[:32]

        # AES-256 decryption
        decrypt_start = time.time()
        aes_data = read_from_file(f'encrypted/ciphertext_{i}.bin')
        nonce_read = aes_data[:16]
        tag_read = aes_data[-16:]
        ciphertext_read = aes_data[16:-16]
        decrypted_message = aes_decrypt(nonce_read, ciphertext_read, tag_read, aes_key_dec)
        decrypt_end = time.time()
        decrypt_time = decrypt_end - decrypt_start

        write_to_file(f'decrypted/decrypted_row_{i}.bin', decrypted_message)

        total_time = time.time() - start_time

        # Log computation times to CSV
        write_to_csv('results/computation_times.csv', [
            i, data_size, encap_time, encrypt_time, decrypt_time, decap_time, total_time
        ])

        if message != decrypted_message:
            print(f"Iteration {i}: Decryption mismatch!")
            continue

        print(f"Iteration {i}: Success.")
        decrypted_rows.append(row)

    # Save decrypted rows to a new XLSX file
    write_rows_to_xlsx("decrypted/decrypted_city.xlsx", decrypted_rows)

    if compare_xlsx_files("city.xlsx", "decrypted/decrypted_city.xlsx"):
        print("Original and decrypted XLSX files match.")
    else:
        print("Mismatch in XLSX file comparison.")
