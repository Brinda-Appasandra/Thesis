import os
import time
import csv
import shutil
from Crypto.Cipher import AES
from Crypto.Hash import SHA256
from Crypto.Random import get_random_bytes
from Crypto.PublicKey import RSA
from Crypto.Signature import pkcs1_15
from openpyxl import load_workbook, Workbook

# Generate RSA key pair
def generate_rsa_keypair():
    key = RSA.generate(2048)
    private_key = key.export_key()
    public_key = key.publickey().export_key()
    return public_key, private_key

# RSA signing
def rsa_sign(message, private_key):
    key = RSA.import_key(private_key)
    h = SHA256.new(message)
    signature = pkcs1_15.new(key).sign(h)
    return signature

# RSA signature verification
def rsa_verify(message, signature, public_key):
    key = RSA.import_key(public_key)
    h = SHA256.new(message)
    try:
        pkcs1_15.new(key).verify(h, signature)
        return True
    except (ValueError, TypeError):
        return False

# AES-128 encryption using EAX mode
def aes_encrypt(plaintext, key):
    cipher = AES.new(key, AES.MODE_EAX)
    nonce = cipher.nonce
    ciphertext, tag = cipher.encrypt_and_digest(plaintext)
    return nonce, ciphertext, tag

# AES-128 decryption using EAX mode
def aes_decrypt(nonce, ciphertext, tag, key):
    cipher = AES.new(key, AES.MODE_EAX, nonce=nonce)
    plaintext = cipher.decrypt_and_verify(ciphertext, tag)
    return plaintext

# Write data to file
def write_to_file(filepath, data):
    with open(filepath, 'wb') as file:
        file.write(data)

# Read data from file
def read_from_file(filepath):
    with open(filepath, 'rb') as file:
        return file.read()

# Write results to CSV file
def write_to_csv(filepath, data):
    with open(filepath, 'a', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(data)

# Read rows from XLSX file
def read_rows_from_xlsx(filepath):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active
    return [row for row in sheet.iter_rows(values_only=True)]

# Write rows to XLSX file
def write_rows_to_xlsx(filepath, rows):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(filepath)

# Compare two XLSX files
def compare_xlsx_files(file1, file2):
    return read_rows_from_xlsx(file1) == read_rows_from_xlsx(file2)

# Create directories if they don't exist
def create_directories():
    directories = ['keys', 'encrypted', 'decrypted', 'signed', 'results']
    for directory in directories:
        if os.path.exists(directory):
            shutil.rmtree(directory)
        os.makedirs(directory, exist_ok=True)

# Example usage
if __name__ == "__main__":
    create_directories()

    # Generate RSA key pair and save keys
    public_key, private_key = generate_rsa_keypair()
    write_to_file('keys/public_key_rsa.pem', public_key)
    write_to_file('keys/private_key_rsa.pem', private_key)
    print("RSA public and private keys written to keys/public_key_rsa.pem and keys/private_key_rsa.pem")

    # Generate AES-128 key and save it
    aes_key = get_random_bytes(16)  # AES-128 key (16 bytes)
    write_to_file('keys/aes_key.bin', aes_key)
    print("AES-128 key generated and saved to keys/aes_key.bin")

    # Read rows from XLSX file
    rows = read_rows_from_xlsx('city.xlsx')

    # Initialize CSV file with headers
    write_to_csv('results/computation_times.csv', ['Iteration', 'Data Size (bytes)', 'Encrypt Time (seconds)', 'Sign Time (seconds)', 'Verify Time (seconds)', 'Decrypt Time (seconds)', 'Total Time (seconds)'])

    decrypted_rows = []

    # Perform the operations and measure time for each row
    for i, row in enumerate(rows):
        message = str(row).encode()  # Convert row to bytes
        data_size = len(message)

        start_time = time.time()

        # AES-128 encryption
        encrypt_start_time = time.time()
        nonce, ciphertext, tag = aes_encrypt(message, aes_key)
        encrypt_end_time = time.time()
        encrypt_time = encrypt_end_time - encrypt_start_time
        write_to_file(f'encrypted/ciphertext_{i}.bin', nonce + ciphertext + tag)
        print(f"Iteration {i}: AES-128 encrypted message saved to 'encrypted/ciphertext_{i}.bin'.")

        # RSA signing
        sign_start_time = time.time()
        signature = rsa_sign(ciphertext, private_key)
        sign_end_time = time.time()
        sign_time = sign_end_time - sign_start_time
        write_to_file(f'signed/signature_{i}.bin', signature)
        print(f"Iteration {i}: RSA signature saved to 'signed/signature_{i}.bin'.")

        # RSA verification
        verify_start_time = time.time()
        signature_read = read_from_file(f'signed/signature_{i}.bin')
        is_verified = rsa_verify(ciphertext, signature_read, public_key)
        verify_end_time = time.time()
        verify_time = verify_end_time - verify_start_time
        print(f"Iteration {i}: Verification {'succeeded' if is_verified else 'failed'} for 'signed/signature_{i}.bin'.")

        # AES-128 decryption
        decrypt_start_time = time.time()
        aes_ciphertext_data = read_from_file(f'encrypted/ciphertext_{i}.bin')
        nonce_read = aes_ciphertext_data[:16]
        tag_read = aes_ciphertext_data[-16:]
        ciphertext_read = aes_ciphertext_data[16:-16]
        decrypted_message = aes_decrypt(nonce_read, ciphertext_read, tag_read, aes_key)
        decrypt_end_time = time.time()
        decrypt_time = decrypt_end_time - decrypt_start_time
        write_to_file(f'decrypted/decrypted_row_{i}.bin', decrypted_message)
        print(f"Iteration {i}: AES-128 decrypted message saved to 'decrypted/decrypted_row_{i}.bin'.")

        end_time = time.time()
        total_time = end_time - start_time

        # Log data size and computation times to CSV
        write_to_csv('results/computation_times.csv', [i, data_size, encrypt_time, sign_time, verify_time, decrypt_time, total_time])
        print(f"Iteration {i}, Data Size: {data_size} bytes, Encrypt Time: {encrypt_time:.6f} sec, Sign Time: {sign_time:.6f} sec, Verify Time: {verify_time:.6f} sec, Decrypt Time: {decrypt_time:.6f} sec, Total Time: {total_time:.6f} sec")

        # Verify the content of the original and decrypted message
        if message != decrypted_message:
            print(f"Iteration {i}: AES-128 decryption failed")
            continue

        # Append decrypted row to list
        decrypted_rows.append(row)

    # Write decrypted rows to a new XLSX file
    write_rows_to_xlsx('decrypted/decrypted_city.xlsx', decrypted_rows)
    print("Decrypted rows written to decrypted/decrypted_city.xlsx")

    # Compare the original and new XLSX files
    if compare_xlsx_files('city.xlsx', 'decrypted/decrypted_city.xlsx'):
        print("The original and decrypted XLSX files are the same.")
    else:
        print("The original and decrypted XLSX files are different.")
